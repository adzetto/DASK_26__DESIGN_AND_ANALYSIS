#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DASK 2026 - Alan Hesabi ve Excel Rapor Olusturucu
V10e modeli yapisal grid verilerinden brut/net alan hesabi.
Profesyonel cok-sayfali Excel ciktisi.
"""

import csv
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, BarChart3D
from pathlib import Path
import os

# ==============================================================================
# 1) VERi OKUMA - V10e pozisyon matrisi
# ==============================================================================
BASE_DIR = Path(r"c:\Users\lenovo\Desktop\DASK_NEW")
CSV_PATH = BASE_DIR / "data" / "twin_position_matrix_v10e.csv"

nodes = []
with open(CSV_PATH, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        nodes.append({
            "id": int(row["node_id"]),
            "x": float(row["x"]),
            "y": float(row["y"]),
            "z": float(row["z"]),
            "floor": int(row["floor"]),
            "zone": row["zone"],
            "tower": row["tower"],
        })

print(f"Toplam dugum: {len(nodes)}")

# ==============================================================================
# 2) GRID ANALIZI
# ==============================================================================
SCALE = 50  # 1:50 olcek

# Tower 1 nodes
t1_nodes = [n for n in nodes if n["tower"] == "1"]
t2_nodes = [n for n in nodes if n["tower"] == "2"]
br_nodes = [n for n in nodes if n["tower"] == "bridge"]

# Unique coordinates
t1_xs = sorted(set(n["x"] for n in t1_nodes))
t1_ys = sorted(set(n["y"] for n in t1_nodes))
t2_ys = sorted(set(n["y"] for n in t2_nodes))

# Floor Z levels
all_z = sorted(set(n["z"] for n in nodes))
floor_z = {i: z for i, z in enumerate(all_z)}  # floor_index -> z

# Bridge Z levels
bridge_z = sorted(set(n["z"] for n in br_nodes))

print(f"Tower 1 X grid: {t1_xs}")
print(f"Tower 1 Y grid: {t1_ys}")
print(f"Tower 2 Y grid: {t2_ys}")
print(f"Kat Z seviyeleri ({len(all_z)} kat): {all_z}")
print(f"Kopru Z seviyeleri: {bridge_z}")

# ==============================================================================
# 3) ALAN HESABI
# ==============================================================================

# --- Convex hull / bounding box per tower per floor ---
def calc_floor_area_bbox(node_list, z_val):
    """Belirli kattaki dugumlerin bounding box alani (cm^2)."""
    floor_nodes = [n for n in node_list if abs(n["z"] - z_val) < 0.01]
    if len(floor_nodes) < 2:
        return 0.0, [], []
    xs = sorted(set(n["x"] for n in floor_nodes))
    ys = sorted(set(n["y"] for n in floor_nodes))
    width_x = max(xs) - min(xs)
    width_y = max(ys) - min(ys)
    return width_x * width_y, xs, ys


def calc_convex_hull_area(node_list, z_val):
    """Convex hull ile kesin alan hesabi (cm^2)."""
    floor_nodes = [n for n in node_list if abs(n["z"] - z_val) < 0.01]
    if len(floor_nodes) < 3:
        xs = [n["x"] for n in floor_nodes]
        ys = [n["y"] for n in floor_nodes]
        if len(floor_nodes) == 2:
            return 0.0, xs, ys
        return 0.0, xs, ys

    points = np.array([[n["x"], n["y"]] for n in floor_nodes])
    # Shoelace formula on convex hull
    from scipy.spatial import ConvexHull
    try:
        hull = ConvexHull(points)
        return hull.volume, points[:, 0].tolist(), points[:, 1].tolist()
    except Exception:
        # Fallback to bounding box
        return calc_floor_area_bbox(node_list, z_val)


# Check scipy availability
try:
    from scipy.spatial import ConvexHull
    USE_HULL = True
    print("scipy.spatial.ConvexHull kullaniliyor (kesin alan).")
except ImportError:
    USE_HULL = False
    print("scipy bulunamadi, bounding box kullaniliyor.")


# --- Kat bazli hesap ---
class FloorData:
    def __init__(self, floor_idx, z_cm):
        self.floor_idx = floor_idx
        self.z_cm = z_cm
        self.z_m = z_cm * SCALE / 100  # gercek olcek metre
        self.h_cm = 0.0  # kat yuksekligi
        self.h_m = 0.0
        # Tower 1
        self.t1_brut_cm2 = 0.0
        self.t1_xs = []
        self.t1_ys = []
        # Tower 2
        self.t2_brut_cm2 = 0.0
        self.t2_xs = []
        self.t2_ys = []
        # Bridge
        self.has_bridge = False
        self.br_brut_cm2 = 0.0
        # Totals
        self.total_brut_cm2 = 0.0
        self.total_brut_m2 = 0.0
        # Core (estimated)
        self.core_cm2 = 0.0
        self.net_cm2 = 0.0
        self.net_m2 = 0.0
        # AutoCAD measured (real scale, m2)
        self.t1_acad_m2 = 0.0
        self.t2_acad_m2 = 0.0
        self.br_acad_m2 = 0.0
        self.total_acad_m2 = 0.0
        self.total_acad_cm2 = 0.0
        self.net_acad_m2 = 0.0
        self.net_acad_cm2 = 0.0
        self.core_m2 = 0.0
        # Usage
        self.usage = ""
        self.plan_type = ""
        self.zone_name = ""


floors = []
for i, z in enumerate(all_z):
    fd = FloorData(i, z)

    # Kat yuksekligi
    if i == 0:
        fd.h_cm = 0  # taban
    elif i == 1:
        fd.h_cm = all_z[1] - all_z[0]  # zemin kat yuksekligi
    else:
        fd.h_cm = all_z[i] - all_z[i - 1]
    fd.h_m = fd.h_cm * SCALE / 100

    # Tower 1 alan
    if USE_HULL:
        fd.t1_brut_cm2, fd.t1_xs, fd.t1_ys = calc_convex_hull_area(t1_nodes, z)
    else:
        fd.t1_brut_cm2, fd.t1_xs, fd.t1_ys = calc_floor_area_bbox(t1_nodes, z)

    # Tower 2 alan
    if USE_HULL:
        fd.t2_brut_cm2, fd.t2_xs, fd.t2_ys = calc_convex_hull_area(t2_nodes, z)
    else:
        fd.t2_brut_cm2, fd.t2_xs, fd.t2_ys = calc_floor_area_bbox(t2_nodes, z)

    # Bridge
    if z in bridge_z:
        fd.has_bridge = True
        br_floor = [n for n in br_nodes if abs(n["z"] - z) < 0.01]
        if len(br_floor) >= 2:
            bxs = [n["x"] for n in br_floor]
            bys = [n["y"] for n in br_floor]
            # Bridge: rectangle from tower1 edge to tower2 edge
            br_width_x = max(bxs) - min(bxs)  # 19-11 = 8 cm
            # Y: tower1 max_y to tower2 min_y
            t1_max_y = max(set(n["y"] for n in t1_nodes if abs(n["z"] - z) < 0.01))
            t2_min_y = min(set(n["y"] for n in t2_nodes if abs(n["z"] - z) < 0.01))
            br_width_y = t2_min_y - t1_max_y  # 24 - 16 = 8 cm
            fd.br_brut_cm2 = br_width_x * br_width_y

    # =====================================================================
    # AutoCAD V2 gercek olcum (DWG gercek olcekte ciziilmis, birim: m)
    # Kat 23 (angular kose): Area=158.6788 m2/kule, Perimeter=55.6697 m
    # Kat 25 (tipik plan):   Area=149.6765 m2/kule, Perimeter=53.7563 m
    # Grid alani: 15x8=120 m2/kule -> doseme kenarlarda grid disina tasiyor
    # =====================================================================
    AUTOCAD_TIPIK_M2 = 149.6765   # Tipik plan per kule (m2, gercek olcek)
    AUTOCAD_ANGULAR_M2 = 158.6788  # Angular kose plani per kule (m2, gercek)
    AUTOCAD_TIPIK_CM2 = AUTOCAD_TIPIK_M2 / (SCALE**2 / 100**2)     # model cm2
    AUTOCAD_ANGULAR_CM2 = AUTOCAD_ANGULAR_M2 / (SCALE**2 / 100**2)  # model cm2

    # Katlara gore gercek doseme alani (AutoCAD olcumu)
    if i in [22, 23, 24]:
        # Angular kose plani (V2 cizimlerinden)
        fd.t1_acad_m2 = AUTOCAD_ANGULAR_M2
        fd.t2_acad_m2 = AUTOCAD_ANGULAR_M2
    elif i == 0:
        fd.t1_acad_m2 = 0
        fd.t2_acad_m2 = 0
    else:
        # Tipik plan (Kat 25 olcumu, = 5,9,13,17,21 ve diger katlar)
        fd.t1_acad_m2 = AUTOCAD_TIPIK_M2
        fd.t2_acad_m2 = AUTOCAD_TIPIK_M2

    fd.br_acad_m2 = fd.br_brut_cm2 * (SCALE**2) / (100**2) if fd.has_bridge else 0
    fd.total_acad_m2 = fd.t1_acad_m2 + fd.t2_acad_m2 + fd.br_acad_m2
    fd.total_acad_cm2 = fd.total_acad_m2 / (SCALE**2 / 100**2)

    # Core tahmini (merdiven + asansor + servis)
    # Goruntulerden: cekirdek ~ 4.4cm x 3.2cm = 14.08 cm2 per tower (model)
    # (merdiven ~2.5x3.2=8, asansor ~1.9x3.2=6.08)
    core_per_tower = 14.08  # cm2
    fd.core_cm2 = core_per_tower * 2  # iki kule
    if fd.has_bridge:
        fd.core_cm2 += 2.0  # kopru koridor/merdiven
    fd.core_m2 = fd.core_cm2 * (SCALE**2) / (100**2)

    # Totals (yapisal grid bazli)
    fd.total_brut_cm2 = fd.t1_brut_cm2 + fd.t2_brut_cm2 + fd.br_brut_cm2
    fd.total_brut_m2 = fd.total_brut_cm2 * (SCALE ** 2) / (100 ** 2)
    fd.net_cm2 = fd.total_brut_cm2 - fd.core_cm2
    fd.net_m2 = fd.net_cm2 * (SCALE ** 2) / (100 ** 2)

    # Net (AutoCAD bazli)
    fd.net_acad_m2 = fd.total_acad_m2 - fd.core_m2
    fd.net_acad_cm2 = fd.net_acad_m2 / (SCALE**2 / 100**2)

    # =====================================================================
    # Kullanim tipleri - IFM Ikiz Kule Mimari Programi
    # V2 AutoCAD cizimleri + Istanbul Finans Merkezi prestij konsepti
    # Kule A: Ofis + Ticaret | Kule B: Ofis (alt) + Rezidans/Otel (ust)
    # Kopruler: H/4, H/2, 3H/4, tepe -> sosyal/gastronomi/sky lounge
    # =====================================================================
    if i == 0:
        fd.usage = "Temel katı (sabit mesnet, betonarme radye temel)"
        fd.plan_type = "-"
        fd.zone_name = "-"
    elif i == 1:
        fd.usage = "Ana giriş lobisi, resepsiyon, güvenlik kontrol, danışma masası, dijital sergi galerisi, ATM/banka şubesi"
        fd.plan_type = "Zemin"
        fd.zone_name = "LOBİ"
    elif i == 2:
        fd.usage = "A: Café & fine-dining restoran, flagship mağaza, co-working lounge | B: Banka operasyon merkezi, müşteri hizmetleri"
        fd.plan_type = "Ticari-1"
        fd.zone_name = "TİCARİ"
    elif i == 3:
        fd.usage = "A: Konferans merkezi (200 kişi), etkinlik salonu, basın odası | B: Eğitim & seminer salonu, atölye alanları"
        fd.plan_type = "Ticari-2"
        fd.zone_name = "TİCARİ"
    elif i == 4:
        fd.usage = "Fitness & spa merkezi, kapalı yüzme havuzu, squash kortu, yoga/pilates stüdyosu, soyunma odaları"
        fd.plan_type = "Spor"
        fd.zone_name = "SOSYAL"
    elif i == 5:
        fd.usage = "Açık plan ofis (yüksek yoğunluk), hot-desk alanları, çalışan kafeteryası, IT destek birimi"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 6:
        # Köprü 1 (H/4): Gastronomi köprüsü
        fd.usage = "KÖPRÜ-1 (GASTRONOMİ): Panoramik restoran, sky bar, açık hava terası, İstanbul Boğazı manzarası"
        fd.plan_type = "Köprü-1"
        fd.zone_name = "KÖPRÜ"
    elif i == 7:
        fd.usage = "Açık plan ofis, toplantı odaları (4-12 kişi), breakout dinlenme alanları, mini mutfak"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 8:
        fd.usage = "Açık plan ofis, sessiz çalışma bölmeleri, telefon kabinleri, odaklanma odaları, arşiv"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 9:
        fd.usage = "Açık plan ofis, 4 köşe yönetici ofisi, ana toplantı salonu, resepsiyon alanı"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 10:
        fd.usage = "Açık plan ofis, proje odaları, beyaz tahta/sunum alanları, çevik çalışma istasyonları"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 11:
        fd.usage = "Açık plan ofis, küçük toplantı birimleri, çevik (agile) çalışma alanı, takım çalışma adaları"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-A"
    elif i == 12:
        # Köprü 2 (H/2): İş dünyası köprüsü
        fd.usage = "KÖPRÜ-2 (İŞ DÜNYASI): VIP toplantı salonu, iş geliştirme merkezi, networking lounge, dijital sunum odası"
        fd.plan_type = "Köprü-2"
        fd.zone_name = "KÖPRÜ"
    elif i == 13:
        fd.usage = "Premium ofis süitleri, bölünebilir açık plan, kiralanabilir bağımsız bölümler, ortak sekreterya"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-B"
    elif i == 14:
        fd.usage = "Premium ofis: hukuk büroları, finans şirketi süit ofisleri, avukat toplantı odaları, kütüphane"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-B"
    elif i == 15:
        fd.usage = "Premium ofis: muhasebe/denetim firmaları, danışmanlık süit ofisleri, müşteri kabul salonu"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-B"
    elif i == 16:
        fd.usage = "Premium ofis: yatırım/portföy yönetimi, fintech süitleri, veri analiz odası, güvenli arşiv"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-B"
    elif i == 17:
        fd.usage = "Üst düzey yönetici katı: CEO süiti, özel sekreterya, VIP bekleme salonu, yönetim kurulu odası"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS-C"
    elif i == 18:
        # Köprü 3 (3H/4): Kültür & sanat köprüsü
        fd.usage = "KÖPRÜ-3 (KÜLTÜR & SANAT): Sanat galerisi, kültür merkezi, sky garden, açık hava heykel bahçesi"
        fd.plan_type = "Köprü-3"
        fd.zone_name = "KÖPRÜ"
    elif i == 19:
        # Kule A: üst düzey ofis | Kule B: butik otel başlangıcı
        fd.usage = "A: Üst düzey yönetici ofisi, özel toplantı | B: 5 yıldızlı butik otel lobisi, concierge, lounge bar"
        fd.plan_type = "Karma"
        fd.zone_name = "OFİS-C / OTEL"
    elif i == 20:
        fd.usage = "A: Üst düzey şirket ofisi, hukuk departmanı | B: Otel standart oda katı (24 oda), oda servisi"
        fd.plan_type = "Karma"
        fd.zone_name = "OFİS-C / OTEL"
    elif i == 21:
        fd.usage = "A: Üst düzey ofis, strateji odası | B: Otel süit oda katı (12 süit), executive lounge"
        fd.plan_type = "Karma"
        fd.zone_name = "OFİS-C / OTEL"
    elif i == 22:
        # V2 çizim: Spa/havuz, angular köşe, yönetici
        fd.usage = "A: Yönetici ofis süiti, özel spa & dinlenme | B: Otel premium süit, wellness merkezi, jakuzi"
        fd.plan_type = "Tip-C"
        fd.zone_name = "PREMİUM"
    elif i == 23:
        # V2 çizim: Büyük toplantı, süit ofis
        fd.usage = "A: CEO süiti, yönetim kurulu salonu (boardroom) | B: Presidential süit, özel yemek salonu, şömine"
        fd.plan_type = "Tip-D"
        fd.zone_name = "PREMİUM"
    elif i == 24:
        # V2 çizim: Rezidans/suite
        fd.usage = "A: Penthouse ofis, VIP kabul salonu, özel teras | B: Penthouse rezidans, kış bahçesi, açık teras"
        fd.plan_type = "Tip-E"
        fd.zone_name = "PENTHOUSE"
    elif i == 25:
        # Köprü 4 (tepe): Sky lounge
        fd.usage = "KÖPRÜ-4 (SKY LOUNGE): 360° panorama bar, helipad, çatı teras restoranı, gözlem platformu"
        fd.plan_type = "Tip-A + Köprü-4"
        fd.zone_name = "SKY LOUNGE"
    else:
        fd.usage = "Ofis"
        fd.plan_type = "Tip-A"
        fd.zone_name = "OFİS"

    floors.append(fd)

# Skip floor 0 (base) for area summaries
active_floors = floors[1:]  # kat 1-25

print(f"\n{'='*90}")
print(f"{'KAT':>4} {'Z(cm)':>7} {'Grid(cm2)':>10} {'AutoCAD(m2)':>12} {'Kopru':>8} {'Grid Tot':>10} {'AcadTot(m2)':>12}")
print(f"{'='*90}")
total_brut_cm2 = 0
total_brut_m2 = 0
total_acad_m2 = 0
for fd in active_floors:
    br_str = f"{fd.br_brut_cm2:.1f}" if fd.has_bridge else "-"
    print(f"{fd.floor_idx:>4} {fd.z_cm:>7.1f} {fd.total_brut_cm2:>10.1f} {fd.t1_acad_m2:>12.2f} "
          f"{br_str:>8} {fd.total_brut_cm2:>10.1f} {fd.total_acad_m2:>12.2f}")
    total_brut_cm2 += fd.total_brut_cm2
    total_brut_m2 += fd.total_brut_m2
    total_acad_m2 += fd.total_acad_m2

print(f"{'='*90}")
print(f"{'TOPLAM':>34} {total_brut_cm2:>28.1f} {total_acad_m2:>12.2f}")
print(f"\nGrid toplam: {total_brut_m2:.1f} m2 | AutoCAD toplam: {total_acad_m2:.2f} m2")
print(f"Fark: +{total_acad_m2 - total_brut_m2:.2f} m2 ({(total_acad_m2/total_brut_m2 - 1)*100:.1f}%)")

# ==============================================================================
# 4) EXCEL OLUSTURMA
# ==============================================================================
print("\nExcel dosyasi olusturuluyor...")

wb = Workbook()

# --- Renkler ---
DARK_NAVY = PatternFill("solid", fgColor="1B2A4A")
NAVY = PatternFill("solid", fgColor="2C3E6B")
HEADER_BLUE = PatternFill("solid", fgColor="3B5998")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
BRIDGE_GOLD = PatternFill("solid", fgColor="FFF2CC")
BRIDGE_HEADER = PatternFill("solid", fgColor="F4B942")
GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")
GREEN_HEADER = PatternFill("solid", fgColor="548235")
RED_LIGHT = PatternFill("solid", fgColor="FCE4EC")
RED_HEADER = PatternFill("solid", fgColor="C0392B")
ORANGE_LIGHT = PatternFill("solid", fgColor="FFF3E0")
GRAY_LIGHT = PatternFill("solid", fgColor="F5F5F5")
GRAY_MEDIUM = PatternFill("solid", fgColor="E0E0E0")
WHITE = PatternFill("solid", fgColor="FFFFFF")
TOTAL_BG = PatternFill("solid", fgColor="1B2A4A")

# Fonts
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="1B2A4A")
DATA_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_DATA = Font(name="Calibri", size=10, bold=True, color="333333")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BRIDGE_FONT = Font(name="Calibri", size=10, bold=True, color="8B6914")
LINK_FONT = Font(name="Calibri", size=10, color="3B5998", underline="single")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
BOTTOM_THICK = Border(
    bottom=Side(style="medium", color="1B2A4A"),
)

# Alignment
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NUM_FMT_CM2 = '#,##0.00'
NUM_FMT_M2 = '#,##0.0'
NUM_FMT_INT = '#,##0'
NUM_FMT_PCT = '0.0%'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title_row(ws, row, title, ncols, fill=DARK_NAVY, font=TITLE_FONT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font
    cell.fill = fill
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 35


def write_header_row(ws, row, headers, fill=HEADER_BLUE, font=HEADER_FONT):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


# ====================== SHEET 1: OZET ======================
ws1 = wb.active
ws1.title = "OZET"
ws1.sheet_properties.tabColor = "1B2A4A"
set_col_widths(ws1, [5, 16, 20, 14, 14, 14, 16, 16, 14, 14])

ncols = 10
write_title_row(ws1, 1, "DASK 2026 - GÜRSOY TOWERS ALAN HESABI (V10e)", ncols)

# Subtitle
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
cell = ws1.cell(row=2, column=1, value="Yapısal Grid: twin_position_matrix_v10e.csv | Ölçek 1:50 | 1680 Düğüm | 25 Kat + Taban")
cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
cell.fill = GRAY_LIGHT
cell.alignment = CENTER
ws1.row_dimensions[2].height = 20

# Headers
headers1 = ["Kat", "Plan Tipi", "Kullanım", "Kule A\n(cm\u00b2)", "Kule B\n(cm\u00b2)", "Köprü\n(cm\u00b2)",
            "Brüt Toplam\n(cm\u00b2)", "Brüt Toplam\n(m\u00b2)", "Net Alan\n(cm\u00b2)", "Net Alan\n(m\u00b2)"]
write_header_row(ws1, 3, headers1)

# Data rows
for idx, fd in enumerate(active_floors):
    r = 4 + idx
    row_fill = BRIDGE_GOLD if fd.has_bridge else (GRAY_LIGHT if idx % 2 == 0 else WHITE)
    data_font = BRIDGE_FONT if fd.has_bridge else DATA_FONT

    vals = [
        fd.floor_idx,
        fd.plan_type,
        fd.usage,
        fd.t1_brut_cm2,
        fd.t2_brut_cm2,
        fd.br_brut_cm2 if fd.has_bridge else 0,
        fd.total_brut_cm2,
        fd.total_brut_m2,
        fd.net_cm2,
        fd.net_m2,
    ]

    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = THIN_BORDER
        if c == 1:
            cell.alignment = CENTER
        elif c in [2, 3]:
            cell.alignment = LEFT
        else:
            cell.alignment = RIGHT
            if c in [4, 5, 6, 7, 9]:
                cell.number_format = NUM_FMT_CM2
            elif c in [8, 10]:
                cell.number_format = NUM_FMT_M2

# --- FORMULAS for totals ---
total_row = 4 + len(active_floors)
ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
cell = ws1.cell(row=total_row, column=1, value="TOPLAM (25 KAT)")
cell.font = TOTAL_FONT
cell.fill = TOTAL_BG
cell.alignment = CENTER

for c in range(4, 11):
    col_letter = get_column_letter(c)
    formula = f"=SUM({col_letter}4:{col_letter}{total_row-1})"
    cell = ws1.cell(row=total_row, column=c, value=formula)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_BG
    cell.border = THIN_BORDER
    cell.alignment = RIGHT
    if c in [4, 5, 6, 7, 9]:
        cell.number_format = NUM_FMT_CM2
    elif c in [8, 10]:
        cell.number_format = NUM_FMT_M2

ws1.row_dimensions[total_row].height = 30

# --- Summary box ---
summ_row = total_row + 2
ws1.merge_cells(start_row=summ_row, start_column=1, end_row=summ_row, end_column=4)
cell = ws1.cell(row=summ_row, column=1, value="GENEL İSTATİSTİKLER")
cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
cell.fill = NAVY
cell.alignment = CENTER

stats = [
    ("Toplam Kat Sayısı", 25, "kat"),
    ("Toplam Düğüm (V10e)", 1680, "adet"),
    ("Kule Plan Boyutu (model)", "30 x 16 cm", ""),
    ("Kule Plan Boyutu (gerçek)", "15 x 8 m", ""),
    ("Zemin Kat Yüksekliği (model/gerçek)", "9 cm / 4.5 m", ""),
    ("Tipik Kat Yüksekliği (model/gerçek)", "6 cm / 3.0 m", ""),
    ("Toplam Yükseklik (model/gerçek)", "153 cm / 76.5 m", ""),
    ("Kuleler Arası Mesafe (model/gerçek)", "8 cm / 4.0 m", ""),
    ("Köprü Katları", "6, 12, 18, 24-25", ""),
    ("Köprü Boyutu (model/gerçek)", "8 x 8 cm / 4 x 4 m", ""),
    ("Ölçek Faktörü", "1:50", ""),
    ("Alan Ölçek Çarpanı", "x 2500", "(50\u00b2)"),
]

for i, (label, val, unit) in enumerate(stats):
    r = summ_row + 1 + i
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws1.cell(row=r, column=1, value=label)
    cell.font = SUBHEADER_FONT
    cell.fill = LIGHT_BLUE if i % 2 == 0 else WHITE
    cell.alignment = LEFT
    cell.border = THIN_BORDER

    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cell = ws1.cell(row=r, column=3, value=f"{val} {unit}".strip())
    cell.font = BOLD_DATA
    cell.fill = LIGHT_BLUE if i % 2 == 0 else WHITE
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# ====================== SHEET 2: MODEL OLCEK (cm) ======================
ws2 = wb.create_sheet("MODEL (cm)")
ws2.sheet_properties.tabColor = "3B5998"
set_col_widths(ws2, [5, 8, 8, 12, 12, 10, 12, 12, 10, 12])

ncols2 = 10
write_title_row(ws2, 1, "MODEL ÖLÇEK ALAN DETAYI (cm)", ncols2)

headers2 = ["Kat", "Z\n(cm)", "h\n(cm)", "Kule A\n(cm\u00b2)", "Kule B\n(cm\u00b2)", "Köprü\n(cm\u00b2)",
            "Brüt\n(cm\u00b2)", "Çekirdek\n(cm\u00b2)", "Verim\n(%)", "Net\n(cm\u00b2)"]
write_header_row(ws2, 2, headers2)

for idx, fd in enumerate(active_floors):
    r = 3 + idx
    row_fill = BRIDGE_GOLD if fd.has_bridge else (GRAY_LIGHT if idx % 2 == 0 else WHITE)

    efficiency = fd.net_cm2 / fd.total_brut_cm2 if fd.total_brut_cm2 > 0 else 0

    vals = [fd.floor_idx, fd.z_cm, fd.h_cm, fd.t1_brut_cm2, fd.t2_brut_cm2,
            fd.br_brut_cm2 if fd.has_bridge else 0,
            fd.total_brut_cm2, fd.core_cm2, efficiency, fd.net_cm2]

    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT if fd.has_bridge else DATA_FONT
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER if c <= 3 else RIGHT
        if c in [4, 5, 6, 7, 8, 10]:
            cell.number_format = NUM_FMT_CM2
        elif c == 9:
            cell.number_format = NUM_FMT_PCT

# Totals with formulas
tr = 3 + len(active_floors)
ws2.cell(row=tr, column=1, value="TOPLAM (25 KAT)").font = TOTAL_FONT
ws2.cell(row=tr, column=1).fill = TOTAL_BG
ws2.cell(row=tr, column=1).alignment = CENTER
for c in [2, 3]:
    ws2.cell(row=tr, column=c).fill = TOTAL_BG
for c in [4, 5, 6, 7, 8, 10]:
    col_l = get_column_letter(c)
    cell = ws2.cell(row=tr, column=c, value=f"=SUM({col_l}3:{col_l}{tr-1})")
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_BG
    cell.alignment = RIGHT
    cell.number_format = NUM_FMT_CM2
    cell.border = THIN_BORDER

# Avg efficiency
cell = ws2.cell(row=tr, column=9, value=f"=AVERAGE(I3:I{tr-1})")
cell.font = TOTAL_FONT
cell.fill = TOTAL_BG
cell.alignment = RIGHT
cell.number_format = NUM_FMT_PCT
cell.border = THIN_BORDER

# ====================== SHEET 3: GERCEK OLCEK (m) ======================
ws3 = wb.create_sheet("GERCEK (m)")
ws3.sheet_properties.tabColor = "548235"
set_col_widths(ws3, [5, 8, 8, 12, 12, 10, 14, 14, 10, 14])

ncols3 = 10
write_title_row(ws3, 1, "GERÇEK ÖLÇEK ALAN DETAYI (m) - 1:50", ncols3, fill=GREEN_HEADER)

headers3 = ["Kat", "Z\n(m)", "h\n(m)", "Kule A\n(m\u00b2)", "Kule B\n(m\u00b2)", "Köprü\n(m\u00b2)",
            "Brüt Toplam\n(m\u00b2)", "Çekirdek\n(m\u00b2)", "Verim\n(%)", "Net Toplam\n(m\u00b2)"]
write_header_row(ws3, 2, headers3, fill=GREEN_HEADER)

for idx, fd in enumerate(active_floors):
    r = 3 + idx
    row_fill = BRIDGE_GOLD if fd.has_bridge else (GREEN_LIGHT if idx % 2 == 0 else WHITE)

    t1_m2 = fd.t1_brut_cm2 * SCALE**2 / 10000
    t2_m2 = fd.t2_brut_cm2 * SCALE**2 / 10000
    br_m2 = fd.br_brut_cm2 * SCALE**2 / 10000 if fd.has_bridge else 0
    core_m2 = fd.core_cm2 * SCALE**2 / 10000
    efficiency = fd.net_cm2 / fd.total_brut_cm2 if fd.total_brut_cm2 > 0 else 0

    vals = [fd.floor_idx, fd.z_m, fd.h_m, t1_m2, t2_m2, br_m2,
            fd.total_brut_m2, core_m2, efficiency, fd.net_m2]

    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT if fd.has_bridge else DATA_FONT
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER if c <= 3 else RIGHT
        if c in [4, 5, 6, 7, 8, 10]:
            cell.number_format = NUM_FMT_M2
        elif c == 9:
            cell.number_format = NUM_FMT_PCT
        elif c in [2, 3]:
            cell.number_format = '0.0'

# Totals
tr = 3 + len(active_floors)
ws3.cell(row=tr, column=1, value="TOPLAM").font = TOTAL_FONT
ws3.cell(row=tr, column=1).fill = GREEN_HEADER
ws3.cell(row=tr, column=1).alignment = CENTER
for c in [2, 3]:
    ws3.cell(row=tr, column=c).fill = GREEN_HEADER
for c in [4, 5, 6, 7, 8, 10]:
    col_l = get_column_letter(c)
    cell = ws3.cell(row=tr, column=c, value=f"=SUM({col_l}3:{col_l}{tr-1})")
    cell.font = TOTAL_FONT
    cell.fill = GREEN_HEADER
    cell.alignment = RIGHT
    cell.number_format = NUM_FMT_M2
    cell.border = THIN_BORDER
cell = ws3.cell(row=tr, column=9, value=f"=AVERAGE(I3:I{tr-1})")
cell.font = TOTAL_FONT
cell.fill = GREEN_HEADER
cell.alignment = RIGHT
cell.number_format = NUM_FMT_PCT
cell.border = THIN_BORDER

# ====================== SHEET 4: KOPRU DETAY ======================
ws4 = wb.create_sheet("KÖPRÜ DETAY")
ws4.sheet_properties.tabColor = "F4B942"
set_col_widths(ws4, [8, 8, 14, 14, 14, 18, 14])

bridge_floors = [fd for fd in floors if fd.has_bridge]
ncols4 = 7
write_title_row(ws4, 1, "KÖPRÜ KATLARI DETAY ANALİZİ", ncols4, fill=BRIDGE_HEADER)

headers4 = ["Kat", "Z (cm)", "Köprü X\n(cm)", "Köprü Y\n(cm)", "Köprü Alan\n(cm\u00b2)",
            "Köprü Alan\n(m\u00b2)", "Kat Toplam\n(cm\u00b2)"]
write_header_row(ws4, 2, headers4, fill=BRIDGE_HEADER)

for idx, fd in enumerate(bridge_floors):
    r = 3 + idx
    # Get bridge dimensions
    br_fl = [n for n in br_nodes if abs(n["z"] - fd.z_cm) < 0.01]
    bxs = [n["x"] for n in br_fl]
    br_x = max(bxs) - min(bxs) if bxs else 0
    t1_max_y = max(set(n["y"] for n in t1_nodes if abs(n["z"] - fd.z_cm) < 0.01), default=16)
    t2_min_y = min(set(n["y"] for n in t2_nodes if abs(n["z"] - fd.z_cm) < 0.01), default=24)
    br_y = t2_min_y - t1_max_y

    vals = [fd.floor_idx, fd.z_cm, br_x, br_y, fd.br_brut_cm2,
            fd.br_brut_cm2 * SCALE**2 / 10000, fd.total_brut_cm2]

    row_fill = BRIDGE_GOLD if idx % 2 == 0 else ORANGE_LIGHT
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if c in [5, 7]:
            cell.number_format = NUM_FMT_CM2
        elif c == 6:
            cell.number_format = NUM_FMT_M2

# Bridge totals
btr = 3 + len(bridge_floors)
ws4.cell(row=btr, column=1, value="TOPLAM").font = TOTAL_FONT
ws4.cell(row=btr, column=1).fill = BRIDGE_HEADER
ws4.cell(row=btr, column=1).alignment = CENTER
for c in [2, 3, 4]:
    ws4.cell(row=btr, column=c).fill = BRIDGE_HEADER
for c in [5, 6, 7]:
    col_l = get_column_letter(c)
    cell = ws4.cell(row=btr, column=c, value=f"=SUM({col_l}3:{col_l}{btr-1})")
    cell.font = TOTAL_FONT
    cell.fill = BRIDGE_HEADER
    cell.alignment = CENTER
    cell.number_format = NUM_FMT_CM2 if c != 6 else NUM_FMT_M2
    cell.border = THIN_BORDER

# ====================== SHEET 5: GRID KOORDINATLARI ======================
ws5 = wb.create_sheet("GRID")
ws5.sheet_properties.tabColor = "C0392B"
set_col_widths(ws5, [8, 14, 14, 14, 14, 14])

write_title_row(ws5, 1, "YAPISAL GRİD KOORDİNATLARI (V10e)", 6, fill=RED_HEADER)

# Tower 1 X grid
ws5.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
cell = ws5.cell(row=3, column=1, value="KULE A - X Koordinatları")
cell.font = SUBHEADER_FONT
cell.fill = LIGHT_BLUE
cell.alignment = CENTER

headers_grid = ["Kolon No", "X (cm)", "X (m)"]
for i, h in enumerate(headers_grid, 1):
    cell = ws5.cell(row=4, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_BLUE
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, x in enumerate(t1_xs):
    r = 5 + i
    ws5.cell(row=r, column=1, value=i + 1).font = DATA_FONT
    ws5.cell(row=r, column=1).alignment = CENTER
    ws5.cell(row=r, column=1).border = THIN_BORDER
    ws5.cell(row=r, column=1).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

    ws5.cell(row=r, column=2, value=x).font = BOLD_DATA
    ws5.cell(row=r, column=2).alignment = CENTER
    ws5.cell(row=r, column=2).border = THIN_BORDER
    ws5.cell(row=r, column=2).number_format = '0.0'
    ws5.cell(row=r, column=2).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

    ws5.cell(row=r, column=3, value=x * SCALE / 100).font = DATA_FONT
    ws5.cell(row=r, column=3).alignment = CENTER
    ws5.cell(row=r, column=3).border = THIN_BORDER
    ws5.cell(row=r, column=3).number_format = '0.00'
    ws5.cell(row=r, column=3).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

# X spans
span_start = 5 + len(t1_xs) + 1
ws5.merge_cells(start_row=span_start, start_column=1, end_row=span_start, end_column=3)
cell = ws5.cell(row=span_start, column=1, value="X Açıklıkları")
cell.font = SUBHEADER_FONT
cell.fill = ORANGE_LIGHT
cell.alignment = CENTER

for i in range(len(t1_xs) - 1):
    r = span_start + 1 + i
    span_cm = t1_xs[i + 1] - t1_xs[i]
    ws5.cell(row=r, column=1, value=f"{i+1}-{i+2}").font = DATA_FONT
    ws5.cell(row=r, column=1).alignment = CENTER
    ws5.cell(row=r, column=1).border = THIN_BORDER
    ws5.cell(row=r, column=2, value=span_cm).font = BOLD_DATA
    ws5.cell(row=r, column=2).alignment = CENTER
    ws5.cell(row=r, column=2).border = THIN_BORDER
    ws5.cell(row=r, column=2).number_format = '0.0'
    ws5.cell(row=r, column=3, value=span_cm * SCALE / 100).font = DATA_FONT
    ws5.cell(row=r, column=3).alignment = CENTER
    ws5.cell(row=r, column=3).border = THIN_BORDER
    ws5.cell(row=r, column=3).number_format = '0.00'

# Tower 1 Y grid (column 4-6)
ws5.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
cell = ws5.cell(row=3, column=4, value="KULE A - Y Koordinatları")
cell.font = SUBHEADER_FONT
cell.fill = LIGHT_BLUE
cell.alignment = CENTER

for i, h in enumerate(["Kolon No", "Y (cm)", "Y (m)"], 4):
    cell = ws5.cell(row=4, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_BLUE
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for i, y in enumerate(t1_ys):
    r = 5 + i
    ws5.cell(row=r, column=4, value=i + 1).font = DATA_FONT
    ws5.cell(row=r, column=4).alignment = CENTER
    ws5.cell(row=r, column=4).border = THIN_BORDER
    ws5.cell(row=r, column=4).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

    ws5.cell(row=r, column=5, value=y).font = BOLD_DATA
    ws5.cell(row=r, column=5).alignment = CENTER
    ws5.cell(row=r, column=5).border = THIN_BORDER
    ws5.cell(row=r, column=5).number_format = '0.0'
    ws5.cell(row=r, column=5).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

    ws5.cell(row=r, column=6, value=y * SCALE / 100).font = DATA_FONT
    ws5.cell(row=r, column=6).alignment = CENTER
    ws5.cell(row=r, column=6).border = THIN_BORDER
    ws5.cell(row=r, column=6).number_format = '0.00'
    ws5.cell(row=r, column=6).fill = GRAY_LIGHT if i % 2 == 0 else WHITE

# ====================== SHEET 6: KAT YUKSEKLIKLERI ======================
ws6 = wb.create_sheet("KAT YÜKSEKLİKLERİ")
ws6.sheet_properties.tabColor = "8E44AD"
set_col_widths(ws6, [5, 10, 10, 10, 10, 10, 10])

write_title_row(ws6, 1, "KAT YÜKSEKLİKLERİ ve KÜMÜLATİF YÜKSEKLİK", 7,
                fill=PatternFill("solid", fgColor="8E44AD"))

headers6 = ["Kat", "Z (cm)", "Z (m)", "h (cm)", "h (m)", "Küm. h (cm)", "Küm. h (m)"]
write_header_row(ws6, 2, headers6, fill=PatternFill("solid", fgColor="8E44AD"))

for idx, fd in enumerate(floors):
    r = 3 + idx
    kum_h = fd.z_cm
    row_fill = BRIDGE_GOLD if fd.has_bridge else (GRAY_LIGHT if idx % 2 == 0 else WHITE)

    vals = [fd.floor_idx, fd.z_cm, fd.z_m, fd.h_cm, fd.h_m, kum_h, kum_h * SCALE / 100]
    for c, v in enumerate(vals, 1):
        cell = ws6.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT if fd.has_bridge else DATA_FONT
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if c in [2, 4, 6]:
            cell.number_format = '0.0'
        elif c in [3, 5, 7]:
            cell.number_format = '0.0'

# ====================== SHEET 7: KULLANIM PLANI ======================
ws7 = wb.create_sheet("KULLANIM PLANI")
ws7.sheet_properties.tabColor = "27AE60"
set_col_widths(ws7, [5, 14, 16, 56, 14, 14, 10])

USAGE_GREEN = PatternFill("solid", fgColor="27AE60")
ncols7 = 7
write_title_row(ws7, 1, "GÜRSOY TOWERS - İFM MİMARİ KULLANIM PROGRAMI", ncols7, fill=USAGE_GREEN)

# Subtitle
ws7.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols7)
cell = ws7.cell(row=2, column=1,
    value="İstanbul Finans Merkezi | Kule A: Ofis+Ticaret | Kule B: Ofis (alt) + Otel/Rezidans (üst) | 4 Köprü: Gastronomi, İş, Kültür, Sky Lounge")
cell.font = Font(name="Calibri", size=9, italic=True, color="2E7D32")
cell.fill = GREEN_LIGHT
cell.alignment = CENTER
ws7.row_dimensions[2].height = 22

headers7 = ["Kat", "Zon", "Plan Tipi", "Kullanım Amacı (Kule A | Kule B)", "AutoCAD\n(m\u00b2)", "Grid\n(m\u00b2)", "Oran\n(%)"]
write_header_row(ws7, 3, headers7, fill=USAGE_GREEN)

# Zone-based color coding (IFM prestij konsepti)
zone_colors = {
    "-":              PatternFill("solid", fgColor="E0E0E0"),   # Temel - gri
    "LOBİ":           PatternFill("solid", fgColor="BBDEFB"),   # Lobi - açık mavi
    "TİCARİ":         PatternFill("solid", fgColor="C8E6C9"),   # Ticari - açık yeşil
    "SOSYAL":         PatternFill("solid", fgColor="B2EBF2"),   # Sosyal/Spor - turkuaz
    "OFİS-A":         PatternFill("solid", fgColor="FFF9C4"),   # Ofis standart - açık sarı
    "OFİS-B":         PatternFill("solid", fgColor="FFE0B2"),   # Ofis premium - açık turuncu
    "OFİS-C":         PatternFill("solid", fgColor="FFCCBC"),   # Ofis üst düzey - somon
    "OFİS-C / OTEL":  PatternFill("solid", fgColor="F3E5F5"),  # Karma - lavanta
    "KÖPRÜ":          BRIDGE_GOLD,                               # Köprü - altın
    "PREMİUM":        PatternFill("solid", fgColor="F8BBD0"),   # Premium - pembe
    "PENTHOUSE":      PatternFill("solid", fgColor="E1BEE7"),   # Penthouse - mor
    "SKY LOUNGE":     PatternFill("solid", fgColor="D1C4E9"),   # Sky lounge - koyu lavanta
    "OTEL":           PatternFill("solid", fgColor="F3E5F5"),   # Otel - açık mor
}

for idx, fd in enumerate(floors):
    r = 4 + idx
    fill = zone_colors.get(fd.zone_name, WHITE)

    pct = fd.total_acad_m2 / total_acad_m2 if total_acad_m2 > 0 else 0
    vals = [fd.floor_idx, fd.zone_name, fd.plan_type, fd.usage,
            fd.total_acad_m2, fd.total_brut_m2, pct]

    for c, v in enumerate(vals, 1):
        cell = ws7.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT if "KÖPRÜ" in str(fd.zone_name) else DATA_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if c in [1, 2]:
            cell.alignment = CENTER
        elif c in [3, 4]:
            cell.alignment = LEFT
        else:
            cell.alignment = RIGHT
        if c == 5:
            cell.number_format = NUM_FMT_M2
        elif c == 6:
            cell.number_format = NUM_FMT_M2
        elif c == 7:
            cell.number_format = NUM_FMT_PCT

# Totals
tr7 = 4 + len(floors)
ws7.merge_cells(start_row=tr7, start_column=1, end_row=tr7, end_column=4)
cell = ws7.cell(row=tr7, column=1, value="TOPLAM (25 KAT)")
cell.font = TOTAL_FONT
cell.fill = USAGE_GREEN
cell.alignment = CENTER
for c in [5, 6]:
    col_l = get_column_letter(c)
    cell = ws7.cell(row=tr7, column=c, value=f"=SUM({col_l}4:{col_l}{tr7-1})")
    cell.font = TOTAL_FONT
    cell.fill = USAGE_GREEN
    cell.alignment = RIGHT
    cell.number_format = NUM_FMT_M2
    cell.border = THIN_BORDER
cell = ws7.cell(row=tr7, column=7, value=f"=SUM(G4:G{tr7-1})")
cell.font = TOTAL_FONT
cell.fill = USAGE_GREEN
cell.alignment = RIGHT
cell.number_format = NUM_FMT_PCT
cell.border = THIN_BORDER

# --- Zone summary table ---
zsr = tr7 + 2
ws7.merge_cells(start_row=zsr, start_column=1, end_row=zsr, end_column=ncols7)
cell = ws7.cell(row=zsr, column=1, value="ZON BAZLI ALAN DAĞITIMI")
cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
cell.fill = USAGE_GREEN
cell.alignment = CENTER

# Aggregate by zone
from collections import OrderedDict
zone_summary = OrderedDict()
for fd in active_floors:
    z = fd.zone_name
    if z not in zone_summary:
        zone_summary[z] = {"acad_m2": 0, "grid_m2": 0, "floors": []}
    zone_summary[z]["acad_m2"] += fd.total_acad_m2
    zone_summary[z]["grid_m2"] += fd.total_brut_m2
    zone_summary[z]["floors"].append(str(fd.floor_idx))

zheaders = ["Zon", "Katlar", "Kat Sayısı", "AutoCAD (m\u00b2)", "Grid (m\u00b2)", "Oran (%)"]
ws7.merge_cells(start_row=zsr+1, start_column=1, end_row=zsr+1, end_column=1)
for ci, h in enumerate(zheaders, 1):
    cell = ws7.cell(row=zsr+1, column=ci, value=h)
    cell.font = HEADER_FONT
    cell.fill = USAGE_GREEN
    cell.alignment = CENTER
    cell.border = THIN_BORDER

grand_acad = sum(v["acad_m2"] for v in zone_summary.values())
for ji, (zname, zdata) in enumerate(zone_summary.items()):
    r = zsr + 2 + ji
    fill = zone_colors.get(zname, WHITE)
    pct = zdata["acad_m2"] / grand_acad if grand_acad > 0 else 0
    floor_list = ",".join(zdata["floors"])
    vals = [zname, floor_list, len(zdata["floors"]), zdata["acad_m2"], zdata["grid_m2"], pct]
    for ci, v in enumerate(vals, 1):
        cell = ws7.cell(row=r, column=ci, value=v)
        cell.font = BOLD_DATA if ci == 1 else DATA_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = CENTER if ci <= 3 else RIGHT
        if ci in [4, 5]:
            cell.number_format = NUM_FMT_M2
        elif ci == 6:
            cell.number_format = NUM_FMT_PCT

# ====================== SHEET 8: AUTOCAD KARSILASTIRMA ======================
ws8 = wb.create_sheet("AUTOCAD ÖLÇÜM")
ws8.sheet_properties.tabColor = "E67E22"
set_col_widths(ws8, [5, 16, 14, 14, 14, 14, 14, 12])

ORANGE_HEADER = PatternFill("solid", fgColor="E67E22")
ncols8 = 8
write_title_row(ws8, 1, "AUTOCAD V2 ÖLÇÜM vs YAPISAL GRİD KARŞILAŞTIRMASI", ncols8, fill=ORANGE_HEADER)

# Subtitle with source info
ws8.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols8)
cell = ws8.cell(row=2, column=1,
    value="Kaynak: 21.01_V2.dwg (gerçek ölçek, m) | Kat 23: 158.68 m\u00b2/kule | Kat 25: 149.68 m\u00b2/kule | Grid: 15x8=120 m\u00b2/kule")
cell.font = Font(name="Calibri", size=9, italic=True, color="8B4513")
cell.fill = ORANGE_LIGHT
cell.alignment = CENTER
ws8.row_dimensions[2].height = 22

headers8 = ["Kat", "Plan Tipi", "Grid\n(m\u00b2/kule)", "AutoCAD\n(m\u00b2/kule)",
            "Kenar Fark\n(m\u00b2/kule)", "Toplam Grid\n(m\u00b2)", "Toplam AutoCAD\n(m\u00b2)", "Fark\n(%)"]
write_header_row(ws8, 3, headers8, fill=ORANGE_HEADER)

for idx, fd in enumerate(active_floors):
    r = 4 + idx
    row_fill = BRIDGE_GOLD if fd.has_bridge else (ORANGE_LIGHT if idx % 2 == 0 else WHITE)

    grid_per_tower = fd.t1_brut_cm2 * SCALE**2 / 10000  # m2
    acad_per_tower = fd.t1_acad_m2
    diff_per_tower = acad_per_tower - grid_per_tower
    pct_diff = diff_per_tower / grid_per_tower if grid_per_tower > 0 else 0

    vals = [fd.floor_idx, fd.plan_type, grid_per_tower, acad_per_tower,
            diff_per_tower, fd.total_brut_m2, fd.total_acad_m2, pct_diff]

    for c, v in enumerate(vals, 1):
        cell = ws8.cell(row=r, column=c, value=v)
        cell.font = BRIDGE_FONT if fd.has_bridge else DATA_FONT
        cell.fill = row_fill
        cell.border = THIN_BORDER
        if c in [1]:
            cell.alignment = CENTER
        elif c == 2:
            cell.alignment = LEFT
        else:
            cell.alignment = RIGHT
        if c in [3, 4, 5, 6, 7]:
            cell.number_format = NUM_FMT_M2
        elif c == 8:
            cell.number_format = '+0.0%;-0.0%'

# Totals
tr8 = 4 + len(active_floors)
ws8.merge_cells(start_row=tr8, start_column=1, end_row=tr8, end_column=2)
cell = ws8.cell(row=tr8, column=1, value="TOPLAM")
cell.font = TOTAL_FONT
cell.fill = ORANGE_HEADER
cell.alignment = CENTER

for c in [3, 4, 5, 6, 7]:
    col_l = get_column_letter(c)
    cell = ws8.cell(row=tr8, column=c, value=f"=SUM({col_l}4:{col_l}{tr8-1})")
    cell.font = TOTAL_FONT
    cell.fill = ORANGE_HEADER
    cell.alignment = RIGHT
    cell.number_format = NUM_FMT_M2
    cell.border = THIN_BORDER

# Avg pct diff
cell = ws8.cell(row=tr8, column=8, value=f"=(G{tr8}-F{tr8})/F{tr8}")
cell.font = TOTAL_FONT
cell.fill = ORANGE_HEADER
cell.alignment = RIGHT
cell.number_format = '+0.0%;-0.0%'
cell.border = THIN_BORDER

# Summary box below
sr = tr8 + 2
summary_data = [
    ("Tipik plan (Kat 25) AutoCAD ölçümü", "149.68 m\u00b2/kule"),
    ("Angular köşe (Kat 23) AutoCAD ölçümü", "158.68 m\u00b2/kule"),
    ("Yapısal grid alanı (15x8)", "120.00 m\u00b2/kule"),
    ("Tipik plan kenar döşeme fazlası", f"+{149.68-120:.2f} m\u00b2/kule (+{(149.68/120-1)*100:.1f}%)"),
    ("Angular köşe kenar döşeme fazlası", f"+{158.68-120:.2f} m\u00b2/kule (+{(158.68/120-1)*100:.1f}%)"),
]
ws8.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
cell = ws8.cell(row=sr, column=1, value="AUTOCAD ÖLÇÜM ÖZETİ")
cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
cell.fill = ORANGE_HEADER
cell.alignment = CENTER

for j, (label, val) in enumerate(summary_data):
    r = sr + 1 + j
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws8.cell(row=r, column=1, value=label)
    cell.font = SUBHEADER_FONT
    cell.fill = ORANGE_LIGHT if j % 2 == 0 else WHITE
    cell.alignment = LEFT
    cell.border = THIN_BORDER

    ws8.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cell = ws8.cell(row=r, column=3, value=val)
    cell.font = BOLD_DATA
    cell.fill = ORANGE_LIGHT if j % 2 == 0 else WHITE
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# ====================== SAVE ======================
output_path = BASE_DIR / "DASK2026_Alan_Hesabi.xlsx"
wb.save(str(output_path))
print(f"\nExcel kaydedildi: {output_path}")
print("Sayfalar: OZET | MODEL (cm) | GERCEK (m) | KOPRU DETAY | GRID | KAT YUKSEKLIKLERI | KULLANIM PLANI | AUTOCAD OLCUM")
